from __future__ import annotations

import re
from pathlib import Path
from urllib.parse import quote_plus

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path(__file__).resolve().parent / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

FORTUNE_SOURCE = (
    "https://raw.githubusercontent.com/EatMoreOranges/"
    "Fortune-500-Dataset/main/data/global-fortune-500-data.csv"
)


def names(text: str) -> list[str]:
    return [line.strip() for line in text.strip().splitlines() if line.strip()]


INDIAN_PRODUCT_CANDIDATES = names(
    """
Flipkart
PhonePe
Razorpay
CRED
Meesho
Swiggy
Zomato
Zepto
Blinkit
Groww
Zerodha
Upstox
Paytm
Policybazaar
Pine Labs
BharatPe
MobiKwik
Jupiter
Fi Money
Navi
Slice
Acko
Digit Insurance
OYO
MakeMyTrip
ixigo
Cleartrip
redBus
Rapido
Ola
Porter
Delhivery
Shiprocket
BlackBuck
Loadshare
Shadowfax
Ecom Express
XpressBees
ElasticRun
Udaan
Moglix
Infra.Market
OfBusiness
Zetwerk
Bizongo
Jumbotail
Ninjacart
DeHaat
WayCool
Captain Fresh
BigBasket
Licious
Country Delight
Rebel Foods
Curefoods
FreshToHome
Urban Company
NoBroker
Housing.com
Magicbricks
PropTiger
Square Yards
Livspace
HomeLane
DesignCafe
Wakefit
Pepperfry
Rentomojo
Furlenco
Cars24
Spinny
CarDekho
Droom
BikeDekho
Ather Energy
Ola Electric
Ultraviolette Automotive
River Mobility
Bounce Infinity
Yulu
Battery Smart
SUN Mobility
Euler Motors
Altigreen
Exponent Energy
ChargeZone
Statiq
Kazam
Log9 Materials
Matter Motor Works
Simple Energy
Raptee.HV
Oben Electric
EMotorad
Noise
boAt
Fire-Boltt
Lenskart
Nykaa
Purplle
Mamaearth
SUGAR Cosmetics
Plum Goodness
Minimalist
The Man Company
Bombay Shaving Company
The Good Glamm Group
Mensa Brands
GlobalBees
FirstCry
Hopscotch
Myntra
AJIO
Bewakoof
The Souled Store
DailyObjects
Chumbak
BlueStone
CaratLane
GIVA
Freshworks
Zoho
Postman
BrowserStack
Chargebee
Druva
Icertis
Innovaccer
Whatfix
Hasura
Uniphore
Yellow.ai
Haptik
Observe.AI
Mad Street Den
SigTuple
Qure.ai
Niramai
Tricog
HealthifyMe
Practo
PharmEasy
Tata 1mg
MediBuddy
mfine
Eka Care
Orange Health Labs
HealthPlix
Cult.fit
Portea
Dozee
BeatO
Ayu Health
Pristyn Care
Scaler
upGrad
Unacademy
Vedantu
Physics Wallah
LEAD Group
Classplus
Teachmint
Cuemath
Testbook
Adda247
Coding Ninjas
InterviewBit
Leverage Edu
CollegeDekho
Sunstone
Masai School
Newton School
NxtWave
Pocket Aces
ShareChat
Dailyhunt
InMobi
Pratilipi
Kuku FM
Pocket FM
Lokal
Josh
Loco
Rooter
Dream11
Games24x7
MPL
WinZO
Nazara Technologies
BetterPlace
Darwinbox
Keka
greytHR
Apna
WorkIndia
Instahyre
Cutshort
HackerEarth
HackerRank
LambdaTest
Cashfree Payments
Open Financial Technologies
KreditBee
Fibe
Yubi
Perfios
Signzy
Zeta
Juspay
Setu
smallcase
Kuvera
INDmoney
CoinDCX
CoinSwitch
Niyo
Simpl
Axio
CredAble
Progcap
M2P Fintech
Hyperface
Decentro
Transak
Airmeet
Exotel
Knowlarity
Amagi
MoEngage
WebEngage
CleverTap
LeadSquared
Capillary Technologies
Wingify
FarMart
AgNext
CropIn
Fasal
Stellapps
AgroStar
Gramophone
BharatAgri
Khatabook
OkCredit
Vyapar
Dukaan
DotPe
GoFrugal
Restroworks
Petpooja
Gupshup
Route Mobile
Unicommerce
Shipway
ClickPost
FarEye
Locus
Shipsy
Fleetx
Vahan
Refyne
Plum Benefits
Onsurity
PazCare
Loop Health
Nova Benefits
Kenko Health
InVideo
Animaker
Vmaker
Rocketlane
Sprinto
Facilio
Zenoti
Mindtickle
HighRadius
Sirion
Unbxd
CleverTap
Media.net
BrowserStack
GreyOrange
Addverb
Ati Motors
Bellatrix Aerospace
Skyroot Aerospace
Agnikul Cosmos
Pixxel
Digantara
Dhruva Space
GalaxEye
ideaForge
Garuda Aerospace
Asteria Aerospace
Tonbo Imaging
Kaleidofin
Jar
StockGro
MarketWolf
Vested Finance
Grip Invest
Wint Wealth
Stable Money
Ditto Insurance
InsuranceDekho
RenewBuy
Turtlemint
Bimaplan
BimaPe
FreightTiger
Wheelseye
Intugine
Detect Technologies
Infinite Uptime
CynLr
Detect Technologies
"""
)

AI_STARTUP_CANDIDATES = names(
    """
OpenAI
Anthropic
Cohere
Mistral AI
xAI
Perplexity
Scale AI
Databricks
Hugging Face
Stability AI
Runway
Midjourney
ElevenLabs
Synthesia
Pika
Luma AI
Suno
Harvey
Sierra
Glean
Writer
Jasper
Copy.ai
Typeface
Adept AI
Imbue
Inflection AI
Character.AI
Replit
Anysphere
Cognition AI
Poolside AI
Magic
Together AI
Fireworks AI
Groq
Cerebras Systems
SambaNova Systems
Lambda
CoreWeave
Crusoe
Modal
Replicate
Baseten
Weights & Biases
Pinecone
Weaviate
Qdrant
Chroma
Zilliz
LangChain
LlamaIndex
Unstructured
Arize AI
Fiddler AI
WhyLabs
Galileo
Patronus AI
Lakera
Protect AI
HiddenLayer
Credo AI
Holistic AI
Arthur AI
Snorkel AI
Labelbox
V7 Labs
Encord
SuperAnnotate
Landing AI
Figure AI
Physical Intelligence
Skild AI
1X Technologies
Sanctuary AI
Agility Robotics
Apptronik
Dexterity
Bright Machines
Nuro
Waabi
Aurora Innovation
Applied Intuition
Gatik
Plus
Pony.ai
May Mobility
Shield AI
Anduril Industries
Helsing
Saronic
Hebbia
Elicit
Consensus
Contextual AI
Decagon
Cresta
Forethought
Parloa
PolyAI
Rasa
Kore.ai
Abridge
Hippocratic AI
Ambience Healthcare
Nabla
Suki
Insitro
Generate Biomedicines
EvolutionaryScale
Owkin
PathAI
Viz.ai
Aidoc
Paige
Freenome
AlphaSense
Captions
Tavus
HeyGen
DeepL
Grammarly
Speechmatics
AssemblyAI
Deepgram
Cartesia
Hume AI
Vapi
Retell AI
Bland AI
LiveKit
Inworld AI
World Labs
Sakana AI
AI21 Labs
Aleph Alpha
Nscale
Lightmatter
Celestial AI
d-Matrix
Etched
SiMa.ai
Tenstorrent
Graphcore
Together Computer
MosaicML
OctoAI
Predibase
Lepton AI
Anyscale
BentoML
RunPod
Vast.ai
Nebius AI
Lambda Labs
Deci AI
Nomic AI
NVIDIA Omniverse Cloud Startups
OpenEvidence
Mercor
Moonvalley
Krea AI
Ideogram
Black Forest Labs
Recraft
Photoroom
Bria AI
Scenario
Leonardo AI
Kaiber
Wonder Dynamics
D-ID
Hour One
Colossyan
Elai.io
Synthflow AI
Regie.ai
Cluely
Lindy
Dust
Relevance AI
Stack AI
Vellum AI
Humanloop
Braintrust Data
Portkey
Helicone
Langfuse
TruEra
Evidently AI
Comet ML
Neptune.ai
Lightning AI
H2O.ai
Dataiku
Domino Data Lab
DataRobot
Datafold
Tecton
Featureform
Feast AI
Cleanlab
Seek AI
Numbers Station
Pecan AI
Obviously AI
Abacus.AI
"""
)

REMOTE_FIRST_CANDIDATES = names(
    """
Automattic
GitLab
Zapier
Buffer
Doist
37signals
Toptal
Remote
Deel
Oyster
SafetyWing
Toggl
Hotjar
Canonical
DuckDuckGo
Wikimedia Foundation
Elastic
Grafana Labs
HashiCorp
Sourcegraph
Sentry
GitBook
Ghost Foundation
Discourse
Help Scout
Kit
Podia
Close
Customer.io
CircleCI
Netlify
Vercel
Supabase
Fly.io
Render
Railway
PlanetScale
Temporal Technologies
Cockroach Labs
Mattermost
Rocket.Chat
Camunda
Airbyte
dbt Labs
Prefect
Dagster Labs
Pulumi
Aiven
Tailscale
1Password
Bitwarden
Proton
Mullvad
Kraken
Coinbase
Chainlink Labs
Consensys
Polygon Labs
Ava Labs
Protocol Labs
Gitcoin
Shopify
Atlassian
HubSpot
Twilio
Dropbox
Reddit
Quora
Stack Overflow
Mozilla
Brave Software
GitHub
DigitalOcean
Cloudflare
Fastly
Algolia
Contentful
Sanity
Storyblok
Strapi
Prismic
Webflow
Framer
Miro
Figma
Sketch
Maze
UserTesting
GitGuardian
Sonar
Snyk
PostHog
Linear
Notion
Coda
Airtable
ClickUp
monday.com
Asana
Loom
Mux
Paddle
Lemon Squeezy
Gumroad
beehiiv
Substack
Patreon
Kajabi
Teachable
Thinkific
Circle
Mighty Networks
Typeform
Jotform
Calendly
Chili Piper
Apollo.io
Clay
Attio
Pipedrive
Intercom
Front
Groove
Kinsta
WP Engine
Cloudways
Semrush
Ahrefs
SE Ranking
Surfer
Omnipresent
Multiplier
Papaya Global
Rippling
Gusto
BambooHR
Lattice
Culture Amp
15Five
Himalayas
FlexJobs
Crossover
Aha!
Articulate
X-Team
BairesDev
Clevertech
10up
Turing
Andela
Contra
Braintrust
Upwork
GitKraken
TestGorilla
Deel Engage
RemotePass
Boundless
WorkMotion
Velocity Global
Multiplier Technologies
Plane
OpenPhone
Dialpad
Aircall
Whereby
Mural
Slite
Twist
Pitch
Pitch Software
Around
Krisp
Kumospace
Gather
Teemyco
Butter
Livestorm
Demio
Hopin
SessionLab
Mentimeter
SaaSFrame
Baremetrics
ChartMogul
ProfitWell
Plausible Analytics
Fathom Analytics
Simple Analytics
Buttondown
Loops
Loops.so
Customerly
Userlist
Churn Buster
Rewardful
FirstPromoter
SavvyCal
Cal.com
Sunsama
Motion
Akiflow
Height
Fibery
Nuclino
Tettra
Slab
ReadMe
Read the Docs
OpenCraft
Platform.sh
Scaleway
Bunny.net
Gcore
Zerodha Tech Remote
Automattic WooCommerce
Canonical Ubuntu
"""
)

INTERNSHIP_HIRER_CANDIDATES = names(
    """
Adobe
NVIDIA
AMD
Qualcomm
Broadcom
Cisco
Oracle
SAP
Salesforce
ServiceNow
Workday
Intuit
Autodesk
Palantir
Snowflake
Datadog
Stripe
Block
PayPal
Visa
Mastercard
American Express
Goldman Sachs
Morgan Stanley
Barclays
UBS
Deutsche Bank
Nomura
Capital One
Bloomberg
Thomson Reuters
Moody's
S&P Global
FactSet
Morningstar
Jane Street
Citadel
Citadel Securities
Two Sigma
Hudson River Trading
Jump Trading
Optiver
IMC Trading
DRW
Akuna Capital
Susquehanna International Group
Tower Research Capital
D. E. Shaw
BlackRock
Vanguard
Fidelity Investments
State Street
BNY Mellon
Northern Trust
KPMG
Deloitte
PwC
EY
Accenture
Capgemini
Cognizant
Infosys
Tata Consultancy Services
Wipro
HCLTech
Tech Mahindra
LTIMindtree
Mphasis
Persistent Systems
ZS Associates
McKinsey & Company
Boston Consulting Group
Bain & Company
Gartner
NielsenIQ
Kearney
Oliver Wyman
Roland Berger
Mercer
Aon
Willis Towers Watson
Schneider Electric
Eaton
Emerson
Rockwell Automation
Caterpillar
Deere & Company
Cummins
Boeing
Airbus
Lockheed Martin
Northrop Grumman
RTX
General Dynamics
SpaceX
Blue Origin
Rocket Lab
Rivian
Lucid Motors
Tesla
Volvo Cars
Uber
Lyft
Airbnb
Booking.com
Expedia Group
Tripadvisor
DoorDash
Instacart
Pinterest
Snap
Spotify
Roblox
Electronic Arts
Epic Games
Riot Games
Unity
Nintendo
Sony Interactive Entertainment
ByteDance
Discord
Twitch
Yelp
Zillow
Redfin
Wayfair
Chewy
Etsy
eBay
Walmart Global Tech
Nike
Adidas
Lululemon
Unilever
Coca-Cola
Mondelez International
Mars
L'Oreal
Merck
Novartis
AstraZeneca
GSK
Sanofi
Eli Lilly
Bristol Myers Squibb
Amgen
Gilead Sciences
Moderna
Regeneron
Biogen
Thermo Fisher Scientific
Danaher
Medtronic
Abbott
Boston Scientific
Stryker
GE HealthCare
Philips
Siemens Healthineers
Splunk
Okta
CrowdStrike
Zscaler
Palo Alto Networks
Fortinet
MongoDB
Confluent
Nutanix
Rubrik
Pure Storage
NetApp
Red Hat
SUSE
Arm
ASML
Applied Materials
Lam Research
KLA
Micron Technology
Texas Instruments
Analog Devices
Marvell Technology
NXP Semiconductors
Infineon Technologies
STMicroelectronics
Synopsys
Cadence Design Systems
Keysight Technologies
MathWorks
Epic Systems
Zoom
DocuSign
Box
Robinhood
SoFi
Chime
Plaid
Brex
Ramp
Affirm
Klarna
Wise
Revolut
N26
Monzo
Starling Bank
Checkout.com
Adyen
Marqeta
Toast
NCR Voyix
FIS
Fiserv
Global Payments
Discover Financial Services
Ally Financial
Charles Schwab
Interactive Brokers
Nasdaq
CME Group
Intercontinental Exchange
London Stock Exchange Group
Deutsche Boerse
Temasek
GIC
Blackstone
KKR
Apollo Global Management
Carlyle
Bain Capital
General Atlantic
Sequoia Capital
Andreessen Horowitz
Accel
Bessemer Venture Partners
Lightspeed Venture Partners
Insight Partners
General Catalyst
Kleiner Perkins
Greylock Partners
Index Ventures
Balderton Capital
Atomico
SoftBank Group
Y Combinator
Techstars
500 Global
Samsung Research
Microsoft Research
Google DeepMind
Allen Institute for AI
Toyota Research Institute
Bosch Research
Honda Research Institute
Disney Research
Dolby Laboratories
Nokia Bell Labs
Ericsson Research
NEC Laboratories
Fujitsu Research
Rakuten
Mercari
Grab
Sea Group
GoTo Group
Gojek
Tokopedia
Shopee
Lazada
Traveloka
Carousell
Naver
Kakao
LINE Yahoo
Rakuten Mobile
BytePlus
Canva
Culture Amp
SafetyCulture
Atlassian Graduate Program
Xero
WiseTech Global
REA Group
Seek
Carsales
Afterpay
Airwallex
Canva Engineering
Thoughtworks
EPAM Systems
Globant
Endava
Luxoft
Publicis Sapient
Slalom
ThoughtSpot
Alteryx
Tableau
MicroStrategy
Qlik
Teradata
Cloudera
C3 AI
UiPath
Automation Anywhere
Celonis
Appian
Pegasystems
Coupa
Veeva Systems
Smartsheet
Procore
Bentley Systems
Trimble
Esri
Garmin
Garmin India
Garmin International
Veritas Technologies
NortonLifeLock
Gen Digital
Akamai Technologies
F5
A10 Networks
Juniper Networks
Arista Networks
Ciena
CommScope
Motorola Solutions
Garmin
Yext
Elastic Path
Contentstack
Acquia
Sitecore
Optimizely
Braze
Amplitude
Mixpanel
AppsFlyer
Adjust
Branch
Iterable
Klaviyo
Sprinklr
Freshworks Global
Zoho Global
Postman Global
BrowserStack Global
Chargebee Global
Druva Global
Icertis Global
Innovaccer Global
Whatfix Global
LambdaTest Global
HackerRank Global
HackerEarth Global
"""
)

# These are discovery candidates only. Their current ATS must be manually verified
# before the pipeline is enabled. The list is intentionally larger than 100 so the
# generated workbook can retain exactly 100 candidates that also exist in Master_1000.
ATS_PROVIDER_CANDIDATES = {
    "OpenAI": "Ashby",
    "Anthropic": "Greenhouse",
    "Cohere": "Greenhouse",
    "Mistral AI": "Lever",
    "Perplexity": "Ashby",
    "Scale AI": "Greenhouse",
    "Hugging Face": "Lever",
    "Runway": "Greenhouse",
    "ElevenLabs": "Ashby",
    "Synthesia": "Greenhouse",
    "Harvey": "Ashby",
    "Sierra": "Ashby",
    "Glean": "Greenhouse",
    "Writer": "Greenhouse",
    "Anysphere": "Ashby",
    "Cognition AI": "Ashby",
    "Together AI": "Greenhouse",
    "Fireworks AI": "Ashby",
    "Groq": "Greenhouse",
    "CoreWeave": "Greenhouse",
    "Modal": "Ashby",
    "Replicate": "Greenhouse",
    "Baseten": "Greenhouse",
    "Weights & Biases": "Greenhouse",
    "Pinecone": "Greenhouse",
    "Weaviate": "Greenhouse",
    "LangChain": "Ashby",
    "LlamaIndex": "Ashby",
    "Arize AI": "Greenhouse",
    "Snorkel AI": "Greenhouse",
    "Labelbox": "Greenhouse",
    "Figure AI": "Greenhouse",
    "Applied Intuition": "Greenhouse",
    "Shield AI": "Greenhouse",
    "Anduril Industries": "Greenhouse",
    "Abridge": "Greenhouse",
    "Hippocratic AI": "Ashby",
    "Deepgram": "Greenhouse",
    "AssemblyAI": "Greenhouse",
    "Vapi": "Ashby",
    "Retell AI": "Ashby",
    "Sakana AI": "Greenhouse",
    "Automattic": "Greenhouse",
    "GitLab": "Greenhouse",
    "Zapier": "Greenhouse",
    "Remote": "Greenhouse",
    "Deel": "Greenhouse",
    "Oyster": "Greenhouse",
    "Canonical": "Greenhouse",
    "Grafana Labs": "Greenhouse",
    "Sourcegraph": "Greenhouse",
    "Sentry": "Greenhouse",
    "Customer.io": "Greenhouse",
    "CircleCI": "Greenhouse",
    "Netlify": "Greenhouse",
    "Vercel": "Ashby",
    "Supabase": "Ashby",
    "Fly.io": "Greenhouse",
    "Render": "Lever",
    "Temporal Technologies": "Greenhouse",
    "Cockroach Labs": "Greenhouse",
    "Airbyte": "Greenhouse",
    "dbt Labs": "Greenhouse",
    "Pulumi": "Greenhouse",
    "Tailscale": "Greenhouse",
    "Coinbase": "Greenhouse",
    "Shopify": "Greenhouse",
    "Atlassian": "Greenhouse",
    "HubSpot": "Greenhouse",
    "Dropbox": "Greenhouse",
    "Reddit": "Greenhouse",
    "Cloudflare": "Greenhouse",
    "Algolia": "Greenhouse",
    "Contentful": "Greenhouse",
    "Webflow": "Greenhouse",
    "Figma": "Greenhouse",
    "Maze": "Greenhouse",
    "Snyk": "Greenhouse",
    "PostHog": "Ashby",
    "Linear": "Ashby",
    "Notion": "Greenhouse",
    "Airtable": "Greenhouse",
    "Loom": "Greenhouse",
    "Paddle": "Greenhouse",
    "Calendly": "Greenhouse",
    "Apollo.io": "Greenhouse",
    "Clay": "Ashby",
    "Attio": "Ashby",
    "Intercom": "Greenhouse",
    "Kinsta": "Greenhouse",
    "Semrush": "Greenhouse",
    "Rippling": "Greenhouse",
    "Lattice": "Greenhouse",
    "Culture Amp": "Greenhouse",
    "TestGorilla": "Greenhouse",
    "Razorpay": "Greenhouse",
    "CRED": "Greenhouse",
    "Meesho": "Greenhouse",
    "Swiggy": "Greenhouse",
    "Zepto": "Ashby",
    "Groww": "Greenhouse",
    "BrowserStack": "Greenhouse",
    "Chargebee": "Greenhouse",
    "Postman": "Greenhouse",
    "Hasura": "Greenhouse",
    "Whatfix": "Greenhouse",
    "Innovaccer": "Greenhouse",
    "Yellow.ai": "Greenhouse",
    "Observe.AI": "Greenhouse",
    "Qure.ai": "Greenhouse",
    "Darwinbox": "Greenhouse",
    "LambdaTest": "Greenhouse",
    "Cashfree Payments": "Greenhouse",
    "Amagi": "Greenhouse",
    "MoEngage": "Greenhouse",
    "CleverTap": "Greenhouse",
    "Rocketlane": "Greenhouse",
    "Sprinto": "Greenhouse",
    "Adobe": "Workday/Other",
    "NVIDIA": "Workday/Other",
    "Datadog": "Greenhouse",
    "Stripe": "Greenhouse",
    "Airbnb": "Greenhouse",
    "Pinterest": "Greenhouse",
    "Snap": "Greenhouse",
    "Spotify": "Greenhouse",
    "Roblox": "Greenhouse",
    "Epic Games": "Greenhouse",
    "Discord": "Greenhouse",
    "MongoDB": "Greenhouse",
    "Confluent": "Greenhouse",
    "Rubrik": "Greenhouse",
    "Okta": "Greenhouse",
    "CrowdStrike": "Greenhouse",
    "Snowflake": "Greenhouse",
    "Palantir": "Lever",
    "Plaid": "Greenhouse",
    "Brex": "Greenhouse",
    "Ramp": "Greenhouse",
    "Chime": "Greenhouse",
    "Robinhood": "Greenhouse",
    "Canva": "Greenhouse",
    "Thoughtworks": "Greenhouse",
    "EPAM Systems": "Greenhouse",
    "Globant": "Greenhouse",
}


def clean_name(value: str) -> str:
    return re.sub(r"\s+", " ", str(value)).strip()


def key(value: str) -> str:
    value = clean_name(value).lower()
    value = value.replace("&", "and")
    return re.sub(r"[^a-z0-9]+", "", value)


def pick_unique(
    candidate_names: list[str], count: int, used: set[str]
) -> list[str]:
    selected: list[str] = []
    for company in candidate_names:
        company = clean_name(company)
        company_key = key(company)
        if not company or company_key in used:
            continue
        used.add(company_key)
        selected.append(company)
        if len(selected) == count:
            return selected
    raise RuntimeError(
        f"Only selected {len(selected)} of required {count}; add more candidates."
    )


def build_row(
    company_id: int,
    company_name: str,
    primary_category: str,
    country: str = "To verify",
    industry: str = "To verify",
    website: str = "",
    source_rank: str = "",
    source_list: str = "Curated JobFight AI seed list",
) -> dict[str, object]:
    high_priority = primary_category != "Big MNC"
    return {
        "Company_ID": f"JF{company_id:04d}",
        "Company_Name": company_name,
        "Primary_Category": primary_category,
        "Country_or_Region": country or "To verify",
        "Industry": industry or "To verify",
        "Official_Website": website or "",
        "Careers_URL": "",
        "Careers_Search_Query": f'"{company_name}" careers jobs',
        "Google_Search_URL": (
            "https://www.google.com/search?q="
            + quote_plus(f'"{company_name}" careers jobs')
        ),
        "ATS_Candidate": "No",
        "ATS_Provider_Candidate": "",
        "ATS_Verification_Status": "Not assessed",
        "Remote_First_Candidate": (
            "Yes" if primary_category == "Remote-first company" else "Unknown"
        ),
        "Recurring_Internship_Hirer_Candidate": (
            "Yes" if primary_category == "Recurring internship hirer" else "Unknown"
        ),
        "AI_Focused_Company": (
            "Yes" if primary_category == "AI startup" else "Unknown"
        ),
        "India_Product_Company": (
            "Yes" if primary_category == "Indian product startup/scale-up" else "No"
        ),
        "Big_MNC": "Yes" if primary_category == "Big MNC" else "No",
        "Monitoring_Priority": "High" if high_priority else "Medium",
        "Recommended_Fetch_Frequency": "Every 6 hours" if high_priority else "Daily",
        "Relevant_Role_Keywords": (
            "AI Intern, ML Intern, GenAI Intern, Data Science Intern, "
            "Software Engineer Intern, Entry-level Engineer"
        ),
        "Active_Hiring_Status": "Unknown - pipeline check required",
        "Last_Careers_Page_Check": "",
        "Current_Openings_Found": "",
        "Source_Rank": source_rank,
        "Source_List": source_list,
        "Record_Verification_Status": (
            "Seed record - verify official careers URL before scraping"
        ),
        "Notes": "",
    }


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="0B1F3A")
    header_font = Font(color="FFFFFF", bold=True)
    accent_fill = PatternFill("solid", fgColor="19A7A0")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        sample_limit = min(ws.max_row, 250)
        for col_idx, cell in enumerate(ws[1], 1):
            values = [str(cell.value or "")]
            for row_idx in range(2, sample_limit + 1):
                values.append(str(ws.cell(row=row_idx, column=col_idx).value or ""))
            width = min(max(len(v) for v in values) + 2, 45)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(width, 12)

        if ws.title == "README":
            for cell in ws[1]:
                cell.fill = accent_fill

    wb.save(path)


def main() -> None:
    fortune = pd.read_csv(FORTUNE_SOURCE)
    fortune = fortune.dropna(subset=["Company"]).head(500).copy()

    used: set[str] = set()
    rows: list[dict[str, object]] = []
    company_id = 1

    # 500 large multinational companies from the public Global Fortune seed CSV.
    for _, record in fortune.iterrows():
        company = clean_name(record.get("Company", ""))
        if not company or key(company) in used:
            continue
        used.add(key(company))
        website = clean_name(record.get("Website", ""))
        if website and not website.startswith(("http://", "https://")):
            website = "https://" + website
        rows.append(
            build_row(
                company_id,
                company,
                "Big MNC",
                country=clean_name(record.get("Country", "")),
                industry=clean_name(record.get("Industry", "")),
                website=website,
                source_rank=str(record.get("Rank", "")),
                source_list="Global Fortune 500 public seed dataset (source vintage: 2022)",
            )
        )
        company_id += 1

    if len(rows) != 500:
        raise RuntimeError(f"Expected 500 MNC rows, generated {len(rows)}")

    category_specs = [
        (
            INDIAN_PRODUCT_CANDIDATES,
            200,
            "Indian product startup/scale-up",
            "India",
            "Product technology / startup",
        ),
        (AI_STARTUP_CANDIDATES, 100, "AI startup", "To verify", "Artificial Intelligence"),
        (
            REMOTE_FIRST_CANDIDATES,
            100,
            "Remote-first company",
            "Global / To verify",
            "Technology / distributed company",
        ),
        (
            INTERNSHIP_HIRER_CANDIDATES,
            100,
            "Recurring internship hirer",
            "Global / To verify",
            "Multiple industries",
        ),
    ]

    for candidate_pool, required, category, country, industry in category_specs:
        selected = pick_unique(candidate_pool, required, used)
        for company in selected:
            rows.append(
                build_row(
                    company_id,
                    company,
                    category,
                    country=country,
                    industry=industry,
                )
            )
            company_id += 1

    df = pd.DataFrame(rows)
    if len(df) != 1000:
        raise RuntimeError(f"Expected 1000 rows, generated {len(df)}")
    if df["Company_Name"].map(key).duplicated().any():
        raise RuntimeError("Duplicate normalized company names found")

    # Keep exactly 100 ATS discovery candidates that are in the master dataset.
    existing = set(df["Company_Name"])
    ats_names = [name for name in ATS_PROVIDER_CANDIDATES if name in existing][:100]
    if len(ats_names) < 100:
        raise RuntimeError(
            f"Only {len(ats_names)} ATS candidates overlap with master data; need 100."
        )
    ats_set = set(ats_names)
    for idx, company in df["Company_Name"].items():
        if company in ats_set:
            df.at[idx, "ATS_Candidate"] = "Yes"
            df.at[idx, "ATS_Provider_Candidate"] = ATS_PROVIDER_CANDIDATES[company]
            df.at[idx, "ATS_Verification_Status"] = (
                "Pending manual verification - do not enable scraper yet"
            )

    # Add deterministic ordering and category rank.
    df["Category_Rank"] = df.groupby("Primary_Category").cumcount() + 1

    csv_path = OUTPUT_DIR / "jobfight_ai_company_seed_1000.csv"
    xlsx_path = OUTPUT_DIR / "jobfight_ai_company_seed_1000.xlsx"
    df.to_csv(csv_path, index=False)

    readme_rows = [
        ["Dataset", "JobFight AI - 1,000 Company Monitoring Seed List"],
        ["Purpose", "Seed registry for discovering and monitoring official company career pages."],
        ["Unique companies", int(df["Company_Name"].nunique())],
        ["Big MNCs", int((df["Primary_Category"] == "Big MNC").sum())],
        ["Indian product startups/scale-ups", int((df["Primary_Category"] == "Indian product startup/scale-up").sum())],
        ["AI startups", int((df["Primary_Category"] == "AI startup").sum())],
        ["Remote-first companies", int((df["Primary_Category"] == "Remote-first company").sum())],
        ["Recurring internship hirers", int((df["Primary_Category"] == "Recurring internship hirer").sum())],
        ["ATS candidates", int((df["ATS_Candidate"] == "Yes").sum())],
        ["Important limitation", "This is a discovery seed list, not a claim that every company is hiring today."],
        ["Careers URL rule", "Fill Careers_URL only after confirming the official company or ATS page."],
        ["ATS rule", "ATS provider values are candidates and must be verified before automated fetching."],
        ["Source vintage", "MNC base uses a public Global Fortune 500 dataset with 2022 fields; revalidate names and websites."],
        ["LinkedIn rule", "Use LinkedIn only for manual discovery; store and publish official application URLs."],
    ]
    readme_df = pd.DataFrame(readme_rows, columns=["Field", "Value"])

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        readme_df.to_excel(writer, sheet_name="README", index=False)
        df.to_excel(writer, sheet_name="Master_1000", index=False)
        for category, sheet in [
            ("Big MNC", "Big_MNCs_500"),
            ("Indian product startup/scale-up", "Indian_Product_200"),
            ("AI startup", "AI_Startups_100"),
            ("Remote-first company", "Remote_First_100"),
            ("Recurring internship hirer", "Internship_Hirers_100"),
        ]:
            df[df["Primary_Category"] == category].to_excel(
                writer, sheet_name=sheet, index=False
            )
        df[df["ATS_Candidate"] == "Yes"].to_excel(
            writer, sheet_name="ATS_Candidates_100", index=False
        )

    style_workbook(xlsx_path)

    print("Generated:")
    print(csv_path)
    print(xlsx_path)
    print(df["Primary_Category"].value_counts().to_dict())
    print("ATS candidates:", int((df["ATS_Candidate"] == "Yes").sum()))


if __name__ == "__main__":
    main()
